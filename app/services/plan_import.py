from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import Select, func, select
from sqlalchemy.orm import Session

from app.models.inventory import EquipmentInstance
from app.models.org import Building, Department, Floor, Room
from app.models.plan import EquipmentCategory, PlanChangeItem, PlanChangeSet, PlanVersion, PlannedItem, PlannedPosition
from app.models.sync import ImportSession
from app.models.enums import (
    CategoryCode,
    ChangeResolutionStatus,
    ChangeSetStatus,
    ChangeType,
    CommunicationsStatus,
    ImportStatus,
    ImportType,
    ItemPresenceStatus,
    PlanVersionStatus,
    PnrStatus,
    SerialState,
)


@dataclass
class ImportResult:
    import_session: ImportSession
    plan_version: PlanVersion
    imported_rooms: int
    imported_positions: int
    imported_items: int
    detected_changes: int = 0
    change_set_id: str | None = None


def _normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _resolve_source_path(raw_path: str) -> Path:
    path = Path(raw_path)
    if path.is_absolute():
        return path
    return (Path.cwd() / path).resolve()


def _get_or_create_building(session: Session, building_code: str, building_name: str) -> Building:
    building = session.scalar(select(Building).where(Building.code == building_code))
    if building is None:
        building = Building(code=building_code, name=building_name, address=None, is_active=True)
        session.add(building)
        session.flush()
    elif building.name != building_name:
        building.name = building_name
    return building


def _get_or_create_category(session: Session, category_code: CategoryCode) -> EquipmentCategory:
    category = session.scalar(select(EquipmentCategory).where(EquipmentCategory.code == category_code))
    if category is None:
        category_name = "Medical equipment" if category_code == CategoryCode.MEDICAL else "Furniture"
        category = EquipmentCategory(code=category_code, name=category_name, is_active=True)
        session.add(category)
        session.flush()
    return category


def _get_or_create_floor(session: Session, building_id, floor_code: str | None) -> Floor | None:
    if not floor_code:
        return None

    floor = session.scalar(
        select(Floor).where(
            Floor.building_id == building_id,
            Floor.code == floor_code,
        )
    )
    if floor is None:
        floor = Floor(building_id=building_id, code=floor_code, name=floor_code, sort_order=0)
        session.add(floor)
        session.flush()
    return floor


def _get_or_create_department(session: Session, building_id, department_name: str | None) -> Department | None:
    if not department_name:
        return None

    department = session.scalar(
        select(Department).where(
            Department.building_id == building_id,
            Department.name == department_name,
        )
    )
    if department is None:
        department = Department(building_id=building_id, name=department_name, sort_order=0)
        session.add(department)
        session.flush()
    return department


def _get_or_create_room(
    session: Session,
    *,
    building_id,
    floor_id,
    department_id,
    room_code: str,
    room_name: str,
) -> Room:
    room = session.scalar(
        select(Room).where(
            Room.building_id == building_id,
            Room.room_code == room_code,
        )
    )
    if room is None:
        room = Room(
            building_id=building_id,
            floor_id=floor_id,
            department_id=department_id,
            room_code=room_code,
            room_name=room_name,
            room_type=None,
            is_active=True,
        )
        session.add(room)
        session.flush()
    else:
        room.floor_id = floor_id
        room.department_id = department_id
        room.room_name = room_name
    return room


def _next_plan_version_no(session: Session, building_id) -> int:
    stmt: Select = select(func.coalesce(func.max(PlanVersion.version_no), 0) + 1).where(PlanVersion.building_id == building_id)
    return int(session.scalar(stmt) or 1)


def _position_key(position: PlannedPosition, room_code: str) -> tuple[str, str, str, str]:
    return (
        position.position_code,
        position.equipment_name,
        position.model_mark or "",
        room_code,
    )


def _position_soft_key(position: PlannedPosition) -> tuple[str, str, str]:
    return (
        position.position_code,
        position.equipment_name,
        position.model_mark or "",
    )


def _build_change_set(
    session: Session,
    *,
    building_id,
    old_plan_version: PlanVersion,
    new_plan_version: PlanVersion,
    created_by_id,
) -> tuple[PlanChangeSet, int]:
    old_positions = session.execute(
        select(PlannedPosition, Room.room_code)
        .join(Room, Room.id == PlannedPosition.room_id)
        .where(PlannedPosition.plan_version_id == old_plan_version.id)
    ).all()
    new_positions = session.execute(
        select(PlannedPosition, Room.room_code)
        .join(Room, Room.id == PlannedPosition.room_id)
        .where(PlannedPosition.plan_version_id == new_plan_version.id)
    ).all()

    old_exact = {_position_key(position, room_code): (position, room_code) for position, room_code in old_positions}
    new_exact = {_position_key(position, room_code): (position, room_code) for position, room_code in new_positions}

    old_soft: dict[tuple[str, str, str], tuple[PlannedPosition, str]] = {
        _position_soft_key(position): (position, room_code) for position, room_code in old_positions
    }

    old_room_codes = {room_code for _, room_code in old_positions}
    new_room_codes = {room_code for _, room_code in new_positions}

    change_set = PlanChangeSet(
        building_id=building_id,
        old_plan_version_id=old_plan_version.id,
        new_plan_version_id=new_plan_version.id,
        created_by=created_by_id,
        status_code=ChangeSetStatus.READY_FOR_REVIEW,
        summary_json={},
    )
    session.add(change_set)
    session.flush()

    detected_changes = 0
    summary_counter: dict[str, int] = {}

    def add_change(change_type: ChangeType, **kwargs) -> None:
        nonlocal detected_changes
        change = PlanChangeItem(
            plan_change_set_id=change_set.id,
            change_type=change_type,
            resolution_status=ChangeResolutionStatus.PENDING,
            **kwargs,
        )
        session.add(change)
        detected_changes += 1
        summary_counter[change_type.value] = summary_counter.get(change_type.value, 0) + 1

    for room_code in sorted(new_room_codes - old_room_codes):
        room = session.scalar(select(Room).where(Room.building_id == building_id, Room.room_code == room_code))
        add_change(ChangeType.ROOM_ADDED, new_room_id=room.id if room else None)

    for room_code in sorted(old_room_codes - new_room_codes):
        room = session.scalar(select(Room).where(Room.building_id == building_id, Room.room_code == room_code))
        add_change(ChangeType.ROOM_REMOVED, old_room_id=room.id if room else None)

    matched_old_exact = set(old_exact.keys()) & set(new_exact.keys())

    for key in matched_old_exact:
        old_position, _ = old_exact[key]
        new_position, _ = new_exact[key]
        if old_position.planned_quantity != new_position.planned_quantity:
            add_change(
                ChangeType.QUANTITY_CHANGED,
                old_planned_position_id=old_position.id,
                new_planned_position_id=new_position.id,
                old_room_id=old_position.room_id,
                new_room_id=new_position.room_id,
                old_payload={"planned_quantity": old_position.planned_quantity},
                new_payload={"planned_quantity": new_position.planned_quantity},
            )

    unmatched_new_exact = [item for key, item in new_exact.items() if key not in matched_old_exact]
    matched_old_soft_keys: set[tuple[str, str, str]] = set()

    for new_position, _ in unmatched_new_exact:
        soft_key = _position_soft_key(new_position)
        old_match = old_soft.get(soft_key)
        if old_match and old_match[0].room_id != new_position.room_id:
            add_change(
                ChangeType.POSITION_MOVED,
                old_planned_position_id=old_match[0].id,
                new_planned_position_id=new_position.id,
                old_room_id=old_match[0].room_id,
                new_room_id=new_position.room_id,
                old_payload={"planned_quantity": old_match[0].planned_quantity},
                new_payload={"planned_quantity": new_position.planned_quantity},
            )
            matched_old_soft_keys.add(soft_key)
        elif not old_match:
            add_change(
                ChangeType.POSITION_ADDED,
                new_planned_position_id=new_position.id,
                new_room_id=new_position.room_id,
                new_payload={"planned_quantity": new_position.planned_quantity},
            )

    for old_position, _ in old_positions:
        exact_key = _position_key(old_position, session.scalar(select(Room.room_code).where(Room.id == old_position.room_id)))
        soft_key = _position_soft_key(old_position)
        if exact_key in matched_old_exact:
            continue
        if soft_key in matched_old_soft_keys:
            continue
        if soft_key not in {_position_soft_key(position) for position, _ in unmatched_new_exact}:
            add_change(
                ChangeType.POSITION_REMOVED,
                old_planned_position_id=old_position.id,
                old_room_id=old_position.room_id,
                old_payload={"planned_quantity": old_position.planned_quantity},
            )

    change_set.summary_json = {
        "old_plan_version_id": str(old_plan_version.id),
        "new_plan_version_id": str(new_plan_version.id),
        "detected_changes": detected_changes,
        "by_type": summary_counter,
    }
    session.flush()
    return change_set, detected_changes


def import_plan_from_excel(
    session: Session,
    *,
    source_path: str,
    building_code: str,
    building_name: str,
    version_label: str,
    category_code: CategoryCode,
    comment_text: str | None,
    created_by_id,
) -> ImportResult:
    resolved_path = _resolve_source_path(source_path)
    if not resolved_path.exists():
        raise FileNotFoundError(f"Source file not found: {resolved_path}")

    workbook = load_workbook(resolved_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    index = {str(header).strip(): idx for idx, header in enumerate(headers) if header is not None}

    required_headers = ["ЭТАЖ", "ОТДЕЛ", "НОМ", "ПОМ", "ПОЗ", "НАИМ", "КОЛ"]
    missing_headers = [header for header in required_headers if header not in index]
    if missing_headers:
        raise ValueError(f"Missing required headers: {', '.join(missing_headers)}")

    building = _get_or_create_building(session, building_code=building_code, building_name=building_name)
    category = _get_or_create_category(session, category_code=category_code)

    next_version_no = _next_plan_version_no(session, building.id)
    current_applied_version = session.scalar(
        select(PlanVersion)
        .where(
            PlanVersion.building_id == building.id,
            PlanVersion.status_code == PlanVersionStatus.APPLIED,
        )
        .order_by(PlanVersion.version_no.desc())
    )

    import_session = ImportSession(
        import_type=ImportType.INITIAL_PLAN if next_version_no == 1 else ImportType.PLAN_UPDATE,
        file_name=resolved_path.name,
        source_path=str(resolved_path),
        status_code=ImportStatus.STARTED,
        started_by=created_by_id,
        summary_json={},
    )
    session.add(import_session)
    session.flush()

    plan_version = PlanVersion(
        building_id=building.id,
        version_no=next_version_no,
        version_label=version_label,
        source_file_name=resolved_path.name,
        status_code=PlanVersionStatus.APPLIED if next_version_no == 1 else PlanVersionStatus.DIFF_READY,
        import_session_id=import_session.id,
        created_by=created_by_id,
        applied_at=datetime.now(timezone.utc) if next_version_no == 1 else None,
        comment_text=comment_text,
    )
    session.add(plan_version)
    session.flush()

    imported_room_codes: set[str] = set()
    imported_positions = 0
    imported_items = 0

    for row_number, row in enumerate(rows, start=2):
        if not any(value not in (None, "") for value in row):
            continue

        floor_code = _normalize_text(row[index["ЭТАЖ"]])
        department_name = _normalize_text(row[index["ОТДЕЛ"]])
        room_code = _normalize_text(row[index["НОМ"]])
        room_name = _normalize_text(row[index["ПОМ"]])
        position_code = _normalize_text(row[index["ПОЗ"]])
        equipment_name = _normalize_text(row[index["НАИМ"]])
        quantity_raw = row[index["КОЛ"]]

        if not room_code or not room_name or not position_code or not equipment_name:
            continue

        try:
            planned_quantity = int(quantity_raw)
        except (TypeError, ValueError):
            continue

        if planned_quantity <= 0:
            continue

        floor = _get_or_create_floor(session, building.id, floor_code)
        department = _get_or_create_department(session, building.id, department_name)
        room = _get_or_create_room(
            session,
            building_id=building.id,
            floor_id=floor.id if floor else None,
            department_id=department.id if department else None,
            room_code=room_code,
            room_name=room_name,
        )

        planned_position = PlannedPosition(
            plan_version_id=plan_version.id,
            import_session_id=import_session.id,
            source_sheet_name=worksheet.title,
            source_row_number=row_number,
            line_no=row[index["No"]] if "No" in index else None,
            building_id=building.id,
            room_id=room.id,
            category_id=category.id,
            position_code=position_code,
            equipment_name=equipment_name,
            description=_normalize_text(row[index["ОПИС"]]) if "ОПИС" in index else None,
            model_mark=_normalize_text(row[index["МАРКА"]]) if "МАРКА" in index else None,
            manufacturer=_normalize_text(row[index["ПОСТ"]]) if "ПОСТ" in index else None,
            planned_quantity=planned_quantity,
            unit_name=_normalize_text(row[index["ЕД"]]) if "ЕД" in index else None,
            mounting_type=_normalize_text(row[index["КРЕП"]]) if "КРЕП" in index else None,
            equipment_type=_normalize_text(row[index["ТИП"]]) if "ТИП" in index else None,
            dimensions_text=_normalize_text(row[index["ГАБ"]]) if "ГАБ" in index else None,
            weight_text=_normalize_text(row[index["ВЕС"]]) if "ВЕС" in index else None,
            notes=_normalize_text(row[index["ПРИМ"]]) if "ПРИМ" in index else None,
            raw_payload={header: row[idx] for header, idx in index.items()},
        )
        session.add(planned_position)
        session.flush()

        for ordinal_no in range(1, planned_quantity + 1):
            item = PlannedItem(
                planned_position_id=planned_position.id,
                room_id=room.id,
                ordinal_no=ordinal_no,
                display_label=f"Экземпляр {ordinal_no}",
                requires_serial=True,
                serial_policy="required_or_not_provided",
                is_active=True,
            )
            session.add(item)
            session.flush()

            equipment_instance = EquipmentInstance(
                planned_item_id=item.id,
                current_room_id=room.id,
                current_storage_zone_id=None,
                current_presence_status=ItemPresenceStatus.NOT_CHECKED,
                serial_number=None,
                serial_state=SerialState.UNKNOWN,
                pnr_status=PnrStatus.NOT_DONE,
                communications_status=CommunicationsStatus.MISSING,
                actual_condition=None,
                completeness_status=None,
                last_check_at=None,
                last_checked_by=None,
                last_event_id=None,
                version_no=1,
            )
            session.add(equipment_instance)
            imported_items += 1

        imported_room_codes.add(room_code)
        imported_positions += 1

    change_set_id: str | None = None
    detected_changes = 0
    if current_applied_version is not None:
        change_set, detected_changes = _build_change_set(
            session,
            building_id=building.id,
            old_plan_version=current_applied_version,
            new_plan_version=plan_version,
            created_by_id=created_by_id,
        )
        change_set_id = str(change_set.id)

    import_session.status_code = ImportStatus.COMPLETED
    import_session.finished_at = datetime.now(timezone.utc)
    import_session.summary_json = {
        "building_code": building.code,
        "building_name": building.name,
        "plan_version_id": str(plan_version.id),
        "plan_version_no": plan_version.version_no,
        "imported_rooms": len(imported_room_codes),
        "imported_positions": imported_positions,
        "imported_items": imported_items,
        "detected_changes": detected_changes,
        "change_set_id": change_set_id,
    }

    session.flush()

    return ImportResult(
        import_session=import_session,
        plan_version=plan_version,
        imported_rooms=len(imported_room_codes),
        imported_positions=imported_positions,
        imported_items=imported_items,
        detected_changes=detected_changes,
        change_set_id=change_set_id,
    )
