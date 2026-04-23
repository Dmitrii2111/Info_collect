from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
SOURCE_PATH = BASE_DIR / "source.xlsx"
TARGET_PATH = BASE_DIR / "source_v2.xlsx"


def main() -> None:
    workbook = load_workbook(SOURCE_PATH)
    sheet = workbook[workbook.sheetnames[0]]

    headers = [cell.value for cell in sheet[1]]
    index = {str(value): idx + 1 for idx, value in enumerate(headers) if value is not None}

    # 1. Rename one room.
    sheet.cell(row=2, column=index["ПОМ"]).value = "Помещение временного хранения чистого белья (обновлено)"

    # 2. Change planned quantity for an existing position.
    sheet.cell(row=3, column=index["КОЛ"]).value = 2

    # 3. Move one position to another room code and room name.
    sheet.cell(row=4, column=index["НОМ"]).value = "1.01.1"
    sheet.cell(row=4, column=index["ПОМ"]).value = "Распаковочная"

    # 4. Add a brand new position row.
    new_row = sheet.max_row + 1
    source_row = 5
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=new_row, column=col).value = sheet.cell(row=source_row, column=col).value

    sheet.cell(row=new_row, column=index["No"]).value = 999999
    sheet.cell(row=new_row, column=index["ЭТАЖ"]).value = "1"
    sheet.cell(row=new_row, column=index["ОТДЕЛ"]).value = "Тестовая зона"
    sheet.cell(row=new_row, column=index["НОМ"]).value = "1.99.1"
    sheet.cell(row=new_row, column=index["ПОМ"]).value = "Тестовое помещение"
    sheet.cell(row=new_row, column=index["ПОЗ"]).value = "TEST.001"
    sheet.cell(row=new_row, column=index["НАИМ"]).value = "Тестовая позиция"
    sheet.cell(row=new_row, column=index["МАРКА"]).value = "MODEL-TEST"
    sheet.cell(row=new_row, column=index["КОЛ"]).value = 3
    sheet.cell(row=new_row, column=index["ПРИМ"]).value = "Добавлено для проверки diff"

    workbook.save(TARGET_PATH)
    print(TARGET_PATH)


if __name__ == "__main__":
    main()
